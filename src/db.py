import os

from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import inspect, text

db = SQLAlchemy()


def ensure_boat_shared_column():
    if not db.engine:
        return

    inspector = inspect(db.engine)
    columns = {column['name'] for column in inspector.get_columns('boats')}
    if 'is_shared' in columns:
        return

    db.session.execute(text('ALTER TABLE boats ADD COLUMN is_shared BOOLEAN'))
    db.session.commit()


def _load_shared_boats_from_excel():
    from openpyxl import load_workbook

    excel_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'boat_list.xlsx'))
    if not os.path.exists(excel_path):
        return []

    try:
        workbook = load_workbook(excel_path, data_only=True)
    except Exception:
        return []

    worksheet = workbook.active
    boats = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 5:
            continue

        city = row[1]
        port = row[2]
        name = row[3]
        url = row[4]
        note = row[5] if len(row) > 5 else None

        if isinstance(city, str):
            city = city.strip()
        if isinstance(port, str):
            port = port.strip()
        if isinstance(name, str):
            name = name.strip()
        if isinstance(url, str):
            url = url.strip()
        if isinstance(note, str):
            note = note.strip()

        if not name or not url:
            continue

        boats.append({
            'name': name,
            'url': url,
            'city': city or '',
            'port': port or '',
            'note': note or '초기 공용 데이터',
        })

    return boats


def _get_app_setting(key: str):
    from models import AppSetting
    setting = AppSetting.query.get(key)
    return setting.value if setting else None


def _set_app_setting(key: str, value: str):
    from models import AppSetting
    setting = AppSetting.query.get(key)
    if setting:
        setting.value = value
    else:
        setting = AppSetting(key=key, value=value)
        db.session.add(setting)
    db.session.commit()


def initialize_shared_boats():
    from models import Boat

    ensure_boat_shared_column()

    if _get_app_setting('shared_boats_initialized') == 'true':
        return

    shared_boats = _load_shared_boats_from_excel()
    if not shared_boats:
        shared_boats = [
            {
                'name': '팀만수',
                'url': 'https://teammansu.kr/index.php?mid=bk',
                'city': '인천',
                'port': '남항(인천항)',
                'note': '초기 공용 데이터',
            },
            {
                'name': '레드헌터',
                'url': 'https://redhunter.sunsang24.com/ship/schedule_fleet',
                'city': '인천',
                'port': '연안부두',
                'note': '초기 공용 데이터',
            },
            {
                'name': '힐링피싱',
                'url': 'https://hl.sunsang24.com/ship/schedule_fleet/202607',
                'city': '안산',
                'port': '오이도항',
                'note': '초기 공용 데이터',
            },
        ]

    for boat_data in shared_boats:
        existing_boat = Boat.query.filter_by(name=boat_data['name']).first()
        if existing_boat:
            existing_boat.url = boat_data['url']
            existing_boat.city = boat_data['city']
            existing_boat.port = boat_data['port']
            existing_boat.note = boat_data['note']
            existing_boat.is_shared = True
            continue

        boat = Boat(
            name=boat_data['name'],
            url=boat_data['url'],
            city=boat_data['city'],
            port=boat_data['port'],
            note=boat_data['note'],
            is_shared=True,
        )
        db.session.add(boat)

    db.session.commit()
    _set_app_setting('shared_boats_initialized', 'true')


def add_boat_instance(name: str, url: str, city: str, port: str, note: str = None, is_shared: bool = True):
    from models import Boat
    boat = Boat(name=name, url=url, city=city, port=port, note=note, is_shared=is_shared)
    db.session.add(boat)
    try:
        db.session.commit()
        return boat
    except Exception:
        db.session.rollback()
        raise

def get_all_boats():
    from models import Boat
    return Boat.query.order_by(Boat.id).all()

def get_boat_by_id(boat_id: int):
    from models import Boat
    return Boat.query.get(boat_id)

# 추가: 배 삭제 함수
def delete_boat(boat_id: int):
    from models import Boat
    boat = Boat.query.get(boat_id)
    if not boat:
        raise ValueError("등록된 배를 찾을 수 없습니다.")
    db.session.delete(boat)
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        raise

def update_boat(boat_id: int, name: str, url: str, city: str, port: str, note: str = None):
    from models import Boat
    boat = Boat.query.get(boat_id)
    if not boat:
        raise ValueError("등록된 배를 찾을 수 없습니다.")
    boat.name = name
    boat.url = url
    boat.city = city
    boat.port = port
    boat.note = note
    try:
        db.session.commit()
        return boat
    except Exception:
        db.session.rollback()
        raise