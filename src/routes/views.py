import io
import os
import openpyxl
from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, current_app, Response, stream_with_context
from flask import send_from_directory
from forms import BoatRegistrationForm, StatusCheckForm, BoatEditForm
from db import add_boat_instance, get_all_boats, delete_boat, get_boat_by_id, update_boat
from services.reservation_checker import check_single_boat
from forms import REGION_CHOICES
from config import CITY_PORT_MAPPING, PORT_COORDINATES, BADA_PORT_IDS
from services.api_response import success_response, error_response, validation_error_response
from services.status_service import StatusPageService
from concurrent.futures import ThreadPoolExecutor, as_completed
import re
import json
import requests

views = Blueprint('views', __name__, template_folder='templates')

@views.route('/')
def index():
    boats = get_all_boats()
    # Boat 객체들을 딕셔너리로 변환하여 JSON 직렬화 가능하게 만듭니다
    boats_dict = [boat.to_dict() for boat in boats]

    # 홈 모달 등록 폼에서 CSRF 를 사용하기 위해 폼 인스턴스를 전달
    form = BoatRegistrationForm()

    return render_template(
        'index.html',
        boats=boats,
        boats_json=boats_dict,
        form=form,
        city_port_map=CITY_PORT_MAPPING
    )

@views.route('/download_excel')
def download_excel():
    boats = get_all_boats()
    
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registered Boats"
    
    # Add header row (비고 포함)
    headers = ["No", "지역", "항구", "등록된 배", "URL", "비고"]
    ws.append(headers)
    
    # Add data rows
    for i, boat in enumerate(boats, start=1):
        row = [i, boat.city, boat.port, boat.name, boat.url, (boat.note or '')]
        ws.append(row)
        
    # Create a virtual file to save the workbook
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create a response
    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=boat_list.xlsx"}
    )

@views.route('/register', methods=['GET', 'POST'])
def register():
    form = BoatRegistrationForm()
    if request.method == 'POST':
        city = request.form.get('city')
        if city in CITY_PORT_MAPPING:
            form.port.choices = [(port, port) for port in CITY_PORT_MAPPING[city]]
    
    if form.validate_on_submit():
        try:
            add_boat_instance(form.name.data, form.url.data, form.city.data, form.port.data, form.note.data)
            flash('배가 등록되었습니다.', 'success')
            return redirect(url_for('views.index'))
        except Exception as e:
            flash(f'등록 중 오류: {e}', 'danger')
    return render_template('register.html', form=form)


def _compute_region_counts(boats):
    """지역별 배 개수 계산"""
    region_sets = {}
    for boat in boats:
        city = getattr(boat, 'city', None) or ''
        registered_name = getattr(boat, 'name', None) or getattr(boat, 'registered_name', None) or ''
        if not city:
            continue
        region_sets.setdefault(city, set()).add(registered_name or '')
    region_counts = {region: len(names) for region, names in region_sets.items()}
    return region_counts, sum(region_counts.values())


def _get_region_boats(boats):
    """지역별 배 목록"""
    region_boats = {}
    for boat in boats:
        city = getattr(boat, 'city', None) or ''
        if not city:
            continue
        region_boats.setdefault(city, []).append({
            'id': boat.id,
            'name': boat.name,
        })
    for city in region_boats:
        region_boats[city].sort(key=lambda x: x['name'])
    return region_boats

@views.route('/edit/<int:boat_id>', methods=['GET', 'POST'])
def edit_boat(boat_id):
    boat = get_boat_by_id(boat_id)
    if not boat:
        flash('해당 배를 찾을 수 없습니다.', 'danger')
        return redirect(url_for('views.index'))

    form = BoatEditForm(obj=boat)
    if request.method == 'POST':
        city = request.form.get('city')
        if city in CITY_PORT_MAPPING:
            form.port.choices = [(port, port) for port in CITY_PORT_MAPPING[city]]

        if form.validate_on_submit():
            try:
                update_boat(boat_id, form.name.data, form.url.data, form.city.data, form.port.data, form.note.data)
                flash('배 정보가 수정되었습니다.', 'success')
                return redirect(url_for('views.index'))
            except Exception as e:
                flash(f'수정 중 오류: {e}', 'danger')
    else:
        # GET 요청 시, 현재 도시의 항구 목록을 설정
        if boat.city in CITY_PORT_MAPPING:
            form.port.choices = [(port, port) for port in CITY_PORT_MAPPING[boat.city]]

    return render_template('edit_boat.html', form=form, boat_id=boat_id)

@views.route('/status', methods=['GET'])
def status():
    """배 예약 현황 조회 페이지
    
    복잡한 로직을 StatusPageService로 분리하여 가독성 향상
    """
    from services.status_service import StatusPageService, DateValidator
    
    form = StatusCheckForm()
    service = StatusPageService()
    
    # 연월일 파라미터 추출
    year, month, day = service.get_date_params_from_request(request)
    
    # 폼에 값 주입
    if year:
        form.year.data = year
    if month:
        form.month.data = month
    if day:
        form.day.data = day
    
    # 기본 페이지 데이터
    region_names = service.get_region_names()
    selected_regions = service.get_selected_regions(request)
    selected_boats = service.get_selected_boats(request)
    
    # 지역별 배 정보 계산
    registered_boats = get_all_boats()
    region_counts, total_registered = _compute_region_counts(registered_boats)
    region_boats = _get_region_boats(registered_boats)
    
    # 날짜가 완전하지 않으면 빈 결과 반환
    if not DateValidator.is_complete(year, month, day):
        context = service.build_render_context(
            form=form,
            entries=[],
            year=year,
            month=month,
            day=day,
            region_names=region_names,
            selected_regions=selected_regions,
            selected_boats=selected_boats,
            region_counts=region_counts,
            total_registered=total_registered,
            region_boats=region_boats,
        )
        return render_template("status.html", **context)
    
    # 날짜 유효성 검증
    is_valid, error_msg = DateValidator.validate(year, month, day)
    if not is_valid:
        flash(error_msg, "warning")
        context = service.build_render_context(
            form=form,
            entries=[],
            year=year,
            month=month,
            day=day,
            region_names=region_names,
            selected_regions=selected_regions,
            selected_boats=selected_boats,
            region_counts=region_counts,
            total_registered=total_registered,
            region_boats=region_boats,
        )
        return render_template("status.html", **context)
    
    # 정상 렌더링 (비동기로 데이터 조회)
    context = service.build_render_context(
        form=form,
        entries=[],
        year=year,
        month=month,
        day=day,
        region_names=region_names,
        selected_regions=selected_regions,
        selected_boats=selected_boats,
        region_counts=region_counts,
        total_registered=total_registered,
        region_boats=region_boats,
    )
    return render_template('status.html', **context)

# API endpoint: 선박별 결과를 NDJSON으로 스트리밍
@views.route('/api/status', methods=['POST'])
def api_status():
    data = request.get_json(silent=True) or request.form
    try:
        year = int(data.get('year'))
        month = int(data.get('month'))
        day = int(data.get('day'))
    except Exception:
        return jsonify({"error": "invalid date"}), 400

    registered_boats = get_all_boats()

    def get_values(name):
        if hasattr(data, 'getlist'):
            return data.getlist(name)
        value = data.get(name)
        if isinstance(value, list):
            return value
        return [value] if value else []

    filter_targets = [region for region in get_values('regions') if region != '전체']
    boats = [boat for boat in registered_boats if boat.city in filter_targets] if filter_targets else registered_boats
    selected_boats = set(get_values('boats'))
    if selected_boats:
        boats = [boat for boat in boats if boat.name in selected_boats]
    debug_enabled = current_app.config.get('DEBUG_LOGGING_ENABLED', False)

    def _persist_snapshots(results, target_date):
        """라이브 조회 결과를 스냅샷 캐시로 남기고, 감시 중인 변화는 알린다.

        실패해도 조용히 넘어간다. 캐시 저장이 안 됐다고 사용자가 방금 본 조회
        결과까지 망치면 안 된다. 다음 조회 때 다시 채워진다.
        """
        from services.snapshot import entries_to_observations
        from services.snapshot_repository import apply_many
        from services.notify.dispatcher import dispatch_all

        try:
            by_boat = {}
            names = {}
            for result in results:
                boat_id = result.get('boat_id')
                if boat_id is None:
                    continue
                observations = entries_to_observations(
                    boat_id, target_date, result.get('entries') or [])
                if observations:
                    by_boat[boat_id] = observations
                    names[boat_id] = result.get('registered_name')

            transitions = apply_many(target_date, by_boat)

            # 이 경로에서도 알림을 보내야 한다. 스냅샷만 갱신하고 넘어가면
            # 스케줄러가 다음에 볼 때는 이미 바뀐 상태라 변화를 못 알아채고,
            # 감시자는 자리가 났는데도 영영 못 받는다.
            # 중복은 Notification 이력이 막는다.
            if transitions:
                dispatch_all(transitions, names)
        except Exception as exc:
            current_app.logger.warning('스냅샷 저장 실패(조회 결과는 정상): %s', exc)

    def process_boat(boat):
        boat_name = getattr(boat, 'name', None) or 'unknown'
        try:
            info = check_single_boat(boat.url, year, month, day, debug_enabled=debug_enabled, known_ship_name=boat_name)
            source_url = info.get('source_url') or boat.url
            entries = [{
                'ship_name': entry.get('ship_name'),
                'status': entry.get('status'),
                'available': entry.get('available'),
                'raw_status_text': entry.get('raw_status_text'),
                'display_status': entry.get('display_status'),
                'source_url': entry.get('used_url') or entry.get('source_url') or entry.get('url') or source_url,
                'url_path': entry.get('used_url_path') or entry.get('url_path') or source_url,
                'fish': entry.get('fish'),
            } for entry in info.get('entries', [])]
            if not entries:
                entries = [{'ship_name': boat_name, 'status': 'unknown', 'available': None,
                            'raw_status_text': '', 'source_url': source_url, 'url_path': source_url, 'fish': None}]
            return {'boat_id': boat.id,
                    'registered_name': boat_name, 'city': boat.city, 'port': boat.port,
                    'query_date': f'{year:04d}-{month:02d}-{day:02d}', 'tide': info.get('tide'),
                    'entries': entries}
        except Exception as exc:
            return {'boat_id': getattr(boat, 'id', None),
                    'registered_name': boat_name, 'city': getattr(boat, 'city', ''),
                    'port': getattr(boat, 'port', ''), 'query_date': f'{year:04d}-{month:02d}-{day:02d}',
                    'tide': None, 'entries': [{'ship_name': boat_name, 'status': 'unknown',
                    'available': None, 'raw_status_text': f'조회 오류: {exc}',
                    'source_url': boat.url, 'url_path': boat.url, 'fish': None}]}

    def stream_results():
        configured_workers = current_app.config.get('STATUS_MAX_WORKERS', 4)
        try:
            max_workers = max(1, int(configured_workers))
        except (TypeError, ValueError):
            max_workers = 4
        max_workers = min(max_workers, len(boats)) if boats else 1

        yield json.dumps({'type': 'start', 'total': len(boats)}, ensure_ascii=False) + '\n'
        completed = 0
        succeeded = set()
        collected = []
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = [executor.submit(process_boat, boat) for boat in boats]
            for future in as_completed(futures):
                result = future.result()
                completed += 1
                succeeded.add(result.get('registered_name'))
                collected.append(result)
                yield json.dumps(result, ensure_ascii=False) + '\n'

        # 어차피 71척을 다 긁었으니 그 결과를 스냅샷으로 남긴다. 다음에 같은
        # 날짜를 열면 이 캐시를 즉시 읽어 조회가 끝난다(현재 라이브 조회 약 114초).
        # 감시 대상만 수집하는 스케줄러와 달리, 이 경로는 화면에 보이는 전부를 채운다.
        #
        # 저장은 워커 스레드가 아니라 여기서 한다. 스레드에는 앱 컨텍스트가 없다.
        # 그리고 배마다 커밋하지 않고 한 번에 쓴다(apply_many) - DB 가 다른
        # 대륙에 있으면 왕복 지연만으로 십수 초가 늘어난다.
        _persist_snapshots(collected, f'{year:04d}-{month:02d}-{day:02d}')

        # 종료 마커. 이 줄이 도착하지 않았다면 스트림이 중간에 잘린 것이다
        # (예: gunicorn --timeout 초과로 워커가 강제 종료). 프론트는 이걸로
        # 완주와 잘림을 구분하고, 못 받은 배만 재조회할 수 있다.
        missing = [b.name for b in boats if b.name not in succeeded]
        yield json.dumps({'type': 'end', 'total': len(boats),
                          'completed': completed, 'missing': missing},
                         ensure_ascii=False) + '\n'

    return Response(
        stream_with_context(stream_results()),
        mimetype='application/x-ndjson',
        headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'},
    )

@views.route('/api/status/cached', methods=['GET'])
def api_status_cached():
    """저장된 스냅샷을 즉시 돌려준다. 라이브 스크래핑을 하지 않는다.

    화면을 열자마자 보여줄 값이다. 라이브 조회는 71척에 약 114초가 걸리는데,
    그 사이 사용자는 빈 화면을 본다. 캐시가 있으면 즉시 채워놓고, 최신이
    필요할 때만 라이브 조회를 누르게 한다.

    값이 언제 확인된 것인지(checked_at)를 함께 내려준다. 신선도를 숨기면
    사용자가 낡은 값을 최신으로 착각한다.
    """
    from datetime import datetime

    from services.snapshot_repository import load_for_dates

    date_str = request.args.get('date')
    if not date_str:
        return jsonify(validation_error_response(
            error='date 파라미터는 필수입니다.',
            message='조회할 날짜를 지정해주세요.')), 400

    try:
        datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        return jsonify(validation_error_response(
            error='date 형식이 잘못되었습니다.',
            message='YYYY-MM-DD 형식이어야 합니다.')), 400

    boats_by_id = {boat.id: boat for boat in get_all_boats()}

    def get_values(name):
        return [v for v in request.args.getlist(name) if v and v != '전체']

    regions = set(get_values('regions'))
    selected = set(get_values('boats'))

    rows = []
    newest = None
    for snap in load_for_dates([date_str]):
        boat = boats_by_id.get(snap.boat_id)
        if boat is None:
            continue                      # 배가 지워졌는데 스냅샷이 남은 경우
        if regions and boat.city not in regions:
            continue
        if selected and boat.name not in selected:
            continue

        if newest is None or (snap.checked_at and snap.checked_at > newest):
            newest = snap.checked_at

        rows.append({
            'boat_id': boat.id,
            'registered_name': boat.name,
            'city': boat.city,
            'port': boat.port,
            'ship_name': snap.ship_name,
            'status': snap.status,
            'available': snap.available,
            'display_status': snap.display_status,
            'fish': snap.fish,
            'source_url': snap.source_url or boat.url,
            'checked_at': snap.checked_at.isoformat() if snap.checked_at else None,
        })

    return jsonify({
        'date': date_str,
        'rows': rows,
        'boat_count': len({r['boat_id'] for r in rows}),
        'total_boats': len(boats_by_id),
        'checked_at': newest.isoformat() if newest else None,
    })


@views.route('/weather')
def weather():
    """날씨 정보 조회 페이지"""
    return render_template('weather.html', 
                         city_port_mapping=CITY_PORT_MAPPING,
                         port_coordinates=PORT_COORDINATES,
                         bada_port_ids=BADA_PORT_IDS)


@views.route('/api/weather', methods=['GET'])
def api_weather():
    """기상청 API를 호출하여 날씨 정보를 가져오는 API
    
    표준화된 응답 포맷 사용
    """
    from datetime import datetime
    
    port = request.args.get('port')
    date_str = request.args.get('date')
    
    if not port or not date_str:
        return jsonify(validation_error_response(
            error='port와 date 파라미터는 필수입니다.',
            message='항구와 날짜를 입력해주세요.'
        )), 400
    
    if port not in PORT_COORDINATES:
        return jsonify(error_response(
            error=f'PORT_NOT_FOUND',
            message=f'{port}의 좌표 정보를 찾을 수 없습니다.'
        )), 404
    
    lat = PORT_COORDINATES[port]['lat']
    lon = PORT_COORDINATES[port]['lon']
    
    try:
        grid = convert_to_grid(lat, lon)
        target_date = datetime.strptime(date_str, '%Y-%m-%d')
        base_date = target_date.strftime('%Y%m%d')
        
        service_key = current_app.config.get('KMA_API_KEY') or os.environ.get('KMA_API_KEY')
        use_sample = not service_key
        
        if use_sample:
            weather_data = generate_sample_weather_data(port, lat, lon)
            return jsonify(success_response(
                data={
                    'lat': lat,
                    'lon': lon,
                    'nx': grid['nx'],
                    'ny': grid['ny'],
                    'data': weather_data,
                    'is_sample': True,
                    'message': '샘플 데이터입니다. 실제 데이터를 보려면 기상청 API 키를 설정해주세요.'
                }
            ))
        
        # 실제 API 호출
        url = 'http://apis.data.go.kr/1360000/VilageFcstInfoService_2.0/getVilageFcst'
        params = {
            'serviceKey': service_key,
            'pageNo': '1',
            'numOfRows': '1000',
            'dataType': 'JSON',
            'base_date': base_date,
            'base_time': '0500',
            'nx': grid['nx'],
            'ny': grid['ny']
        }
        
        response = requests.get(url, params=params, timeout=10)
        
        if response.status_code != 200:
            current_app.logger.warning(f"KMA API call failed with status {response.status_code}")
            weather_data = generate_sample_weather_data(port, lat, lon)
            return jsonify(success_response(
                data={
                    'lat': lat,
                    'lon': lon,
                    'nx': grid['nx'],
                    'ny': grid['ny'],
                    'data': weather_data,
                    'is_sample': True,
                    'message': 'API 호출 실패로 샘플 데이터를 표시합니다.'
                }
            ))
        
        result = response.json()
        weather_data = process_kma_weather_data(result, base_date)
        
        if not weather_data:
            weather_data = generate_sample_weather_data(port, lat, lon)
            return jsonify(success_response(
                data={
                    'lat': lat,
                    'lon': lon,
                    'nx': grid['nx'],
                    'ny': grid['ny'],
                    'data': weather_data,
                    'is_sample': True,
                    'message': '해당 날짜의 실제 데이터가 없어 샘플 데이터를 표시합니다.'
                }
            ))
        
        return jsonify(success_response(
            data={
                'lat': lat,
                'lon': lon,
                'nx': grid['nx'],
                'ny': grid['ny'],
                'data': weather_data,
                'is_sample': False
            }
        ))
        
    except Exception as e:
        current_app.logger.error(f"Weather API error: {e}")
        # 에러 발생 시에도 샘플 데이터 제공
        try:
            weather_data = generate_sample_weather_data(port, lat, lon)
            return jsonify({
                'lat': lat,
                'lon': lon,
                'nx': grid['nx'] if 'grid' in locals() else 0,
                'ny': grid['ny'] if 'grid' in locals() else 0,
                'data': weather_data,
                'error': f'에러가 발생하여 샘플 데이터를 표시합니다: {str(e)}'
            })
        except:
            return jsonify({'error': f'날씨 정보를 가져올 수 없습니다: {str(e)}'}), 500

def convert_to_grid(lat, lon):
    """위경도를 기상청 격자 좌표로 변환"""
    import math
    
    RE = 6371.00877  # 지구 반경(km)
    GRID = 5.0  # 격자 간격(km)
    SLAT1 = 30.0  # 표준위도1
    SLAT2 = 60.0  # 표준위도2
    OLON = 126.0  # 기준점 경도
    OLAT = 38.0  # 기준점 위도
    XO = 43  # 기준점 X좌표
    YO = 136  # 기준점 Y좌표

    DEGRAD = math.pi / 180.0
    re = RE / GRID
    slat1 = SLAT1 * DEGRAD
    slat2 = SLAT2 * DEGRAD
    olon = OLON * DEGRAD
    olat = OLAT * DEGRAD

    sn = math.tan(math.pi * 0.25 + slat2 * 0.5) / math.tan(math.pi * 0.25 + slat1 * 0.5)
    sn = math.log(math.cos(slat1) / math.cos(slat2)) / math.log(sn)
    sf = math.tan(math.pi * 0.25 + slat1 * 0.5)
    sf = math.pow(sf, sn) * math.cos(slat1) / sn
    ro = math.tan(math.pi * 0.25 + olat * 0.5)
    ro = re * sf / math.pow(ro, sn)

    ra = math.tan(math.pi * 0.25 + lat * DEGRAD * 0.5)
    ra = re * sf / math.pow(ra, sn)
    theta = lon * DEGRAD - olon
    if theta > math.pi:
        theta -= 2.0 * math.pi
    if theta < -math.pi:
        theta += 2.0 * math.pi
    theta *= sn

    nx = int(ra * math.sin(theta) + XO + 0.5)
    ny = int(ro - ra * math.cos(theta) + YO + 0.5)

    return {'nx': nx, 'ny': ny}

def process_kma_weather_data(result, base_date):
    """기상청 API 응답 데이터 처리"""
    if not result.get('response', {}).get('body', {}).get('items', {}).get('item'):
        return []
    
    items = result['response']['body']['items']['item']
    time_data = {}
    
    # 시간대별로 데이터 그룹화
    for item in items:
        fcst_date = item['fcstDate']
        fcst_time = item['fcstTime']
        category = item['category']
        value = item['fcstValue']
        
        if fcst_date == base_date:
            time_key = f"{fcst_time[:2]}시"
            if time_key not in time_data:
                time_data[time_key] = {}
            time_data[time_key][category] = value
    
    # 시간대별 데이터를 배열로 변환
    weather_array = []
    for time in sorted(time_data.keys()):
        data = time_data[time]
        
        # 풍향 변환
        wind_dir_deg = float(data.get('VEC', 0))
        direction = get_wind_direction(wind_dir_deg)
        
        # 날씨 아이콘 결정
        sky = data.get('SKY', '1')
        pty = data.get('PTY', '0')
        weather = get_weather_icon(sky, pty)
        
        weather_array.append({
            'time': time,
            'direction': direction,
            'windSpeed': float(data.get('WSD', 0)),
            'maxWindSpeed': float(data.get('WSD', 0)) * 1.5,
            'weather': weather,
            'temp': float(data.get('TMP', 0)),
            'waveHeight': 0.6,  # 기본값
            'wavePeriod': 7.0   # 기본값
        })
    
    return weather_array

def get_wind_direction(deg):
    """풍향 각도를 방위로 변환"""
    dirs = ['북', '북북동', '북동', '동북동', '동', '동남동', '남동', '남남동',
            '남', '남남서', '남서', '서남서', '서', '서북서', '북서', '북북서']
    idx = int((deg + 22.5 * 0.5) / 22.5) % 16
    return dirs[idx]

def get_weather_icon(sky, pty):
    """하늘 상태와 강수 형태로 날씨 아이콘 결정"""
    if pty == '1' or pty == '4':
        return '🌧️'  # 비
    if pty == '2':
        return '🌨️'  # 비/눈
    if pty == '3':
        return '❄️'  # 눈
    if sky == '1':
        return '☀️'  # 맑음
    if sky == '3':
        return '⛅'  # 구름많음
    if sky == '4':
        return '☁️'  # 흐림
    return '🌤️'

def generate_sample_weather_data(port_name, lat, lon):
    """항구별로 다른 샘플 날씨 데이터 생성"""
    import random
    
    # 항구 이름을 시드로 사용하여 일관된 랜덤 값 생성
    seed = hash(port_name) % 10000
    random.seed(seed)
    
    times = ['00시', '03시', '06시', '09시', '12시', '15시', '18시', '21시']
    
    # 위도에 따라 기온 범위 조정 (남쪽이 더 따뜻함)
    base_temp = 15 + (37.5 - lat) * 0.5  # 위도가 낮을수록 기온 높음
    
    # 경도와 위도로 풍향 경향 결정
    wind_direction_base = int((lon - 126) * 10 + (lat - 35) * 5) % 360
    
    data = []
    for i, time in enumerate(times):
        # 시간대별 기온 변화
        hour = int(time.replace('시', ''))
        temp_variation = -3 if hour < 6 else (5 if 12 <= hour < 15 else 0)
        temp = round(base_temp + temp_variation + random.uniform(-2, 2), 1)
        
        # 풍향 (항구별로 다르게)
        wind_deg = (wind_direction_base + random.randint(-30, 30)) % 360
        direction = get_wind_direction(wind_deg)
        
        # 풍속 (연안 지역 특성)
        wind_speed = round(random.uniform(1.5, 6.0), 1)
        max_wind_speed = round(wind_speed * random.uniform(1.3, 1.8), 1)
        
        # 날씨 (일부 랜덤)
        weather_options = ['☀️', '🌤️', '⛅', '☁️']
        if random.random() < 0.15:  # 15% 확률로 비
            weather_options = ['🌧️', '🌦️']
        weather = random.choice(weather_options)
        
        # 파고 (풍속과 연관)
        wave_height = round(wind_speed * 0.15 + random.uniform(0.3, 0.8), 1)
        wave_period = round(random.uniform(4.0, 9.0), 1)
        
        data.append({
            'time': time,
            'direction': direction,
            'windSpeed': wind_speed,
            'maxWindSpeed': max_wind_speed,
            'weather': weather,
            'temp': int(temp),
            'waveHeight': wave_height,
            'wavePeriod': wave_period
        })
    
    return data

# ---------------- Tide (Badatime) Integration -----------------
@views.route('/api/fishing_index')
def api_fishing_index():
    """항구 + 날짜로 KHOA 바다낚시지수를 조회한다. Phase E.

    이 API 는 오늘부터 +5일만 예보한다(실측, 문서에 없음). 그래서 대부분의
    예약 조회 날짜(보통 몇 주 뒤)에는 데이터가 없다. 그건 오류가 아니라
    정상이므로 200 과 available=false 로 응답한다 - 프론트가 자리 안내처럼
    '예보 범위 밖입니다' 를 조용히 보여주면 된다.
    """
    from services.tide import khoa_fishing

    port = request.args.get('port')
    date_str = request.args.get('date')
    if not port or not date_str:
        return jsonify(validation_error_response(
            error='port와 date 파라미터는 필수입니다.',
            message='항구와 날짜를 입력해주세요.')), 400

    if port not in PORT_COORDINATES:
        return jsonify(error_response(
            error='PORT_NOT_FOUND',
            message=f'{port}의 좌표 정보를 찾을 수 없습니다.')), 404

    if not os.environ.get(khoa_fishing.ENV_KEY):
        return jsonify(error_response(
            error='NOT_CONFIGURED',
            message='낚시지수 API 키가 설정되지 않았습니다.')), 503

    coords = PORT_COORDINATES[port]
    try:
        result = khoa_fishing.for_port(coords['lat'], coords['lon'], date_str)
    except Exception as exc:
        current_app.logger.warning('낚시지수 조회 실패: %s', exc)
        return jsonify(error_response(
            error='UPSTREAM_ERROR',
            message='낚시지수 조회 중 오류가 발생했습니다.')), 502

    return jsonify(success_response(data=result))


@views.route('/api/tide')
def api_tide():
    """바다타임 특정 항구 번호(port_id)의 주간(week_container) 정보를 파싱하여 시간대별 데이터 반환.
    요청: /api/tide?port_id=118
    반환 필드: time, wind_dir, wind_speed, weather, temperature, wave_info
    바다타임 페이지에 풍향/풍속/날씨/기온/파고가 모두 없을 수 있으므로 가용한 정보만 구성하고 나머지는 추정/빈값 처리.
    """
    import requests
    from bs4 import BeautifulSoup
    from services.badatime_parser import TideTableParser
    
    port_id = request.args.get('port_id', type=int)
    if not port_id:
        return validation_error_response('port_id 파라미터가 필요합니다.'), 400

    date_str = request.args.get('date')
    base_url = f"https://www.badatime.com/{port_id}/tide"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0 Safari/537.36'
    }
    
    try:
        used_url = f"{base_url}/{date_str}" if date_str else base_url
        resp = requests.get(used_url, headers=headers, timeout=10)
        if resp.status_code != 200:
            return error_response(f'페이지 응답 오류: {resp.status_code}', error_code='HTTP_ERROR'), 502
    except requests.RequestException as e:
        return error_response(f'요청 실패: {str(e)}', error_code='REQUEST_FAILED'), 500

    soup = BeautifulSoup(resp.text, 'html.parser')
    week_container = soup.select_one('.week_container')
    if not week_container:
        return error_response('week_container를 찾을 수 없습니다.', error_code='PARSING_ERROR'), 500

    table = week_container.select_one('table.week_table')
    if not table:
        return error_response('week_table을 찾을 수 없습니다.', error_code='PARSING_ERROR'), 500

    try:
        parser = TideTableParser(table)
        data_out = parser.parse()
        if data_out is None:
            return error_response('테이블 파싱 실패', error_code='PARSING_ERROR'), 500
        
        return success_response({
            'port_id': port_id,
            'source_url': used_url if date_str else base_url,
            'data': data_out,
            'date': date_str
        })
    except Exception as e:
        return error_response(f'파싱 오류: {str(e)}', error_code='PARSING_ERROR'), 500

# New: Parse Badatime graph page and return only summary table + chart script
@views.route('/api/tide_graph', methods=['GET'])
def api_tide_graph():
    """Badatime 그래프 페이지(/{port_id}/graph/{date})에서 요약 테이블(pc_txt_view)과
    차트 컨테이너(#chartdiv) 및 해당 스크립트만 추출해서 반환.
    응답: { status, data: { pc_html, chart_html, script, source_url } }
    """
    import requests
    from bs4 import BeautifulSoup
    import re

    port_id = request.args.get('port_id', type=int)
    date_str = request.args.get('date', default='')
    if not port_id or not date_str:
        return validation_error_response('port_id와 date가 필요합니다.'), 400

    source_url = f"https://www.badatime.com/{port_id}/graph/{date_str}"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0 Safari/537.36'
    }
    
    try:
        resp = requests.get(source_url, headers=headers, timeout=10)
        if resp.status_code != 200:
            return error_response(f'페이지 응답 오류: {resp.status_code}', error_code='HTTP_ERROR'), 502
    except requests.RequestException as e:
        return error_response(f'요청 실패: {str(e)}', error_code='REQUEST_FAILED'), 500

    soup = BeautifulSoup(resp.text, 'html.parser')

    # PC 요약 테이블
    pc_view = soup.select_one('div.pc_txt_view')
    pc_html = pc_view.decode() if pc_view else ''

    # 모바일 요약 테이블
    mo_view = soup.select_one('div.mo_txt_view')
    mo_html = mo_view.decode() if mo_view else ''

    # 차트 컨테이너와 스크립트
    chart_div = soup.select_one('#chartdiv') or soup.select_one('.graph-wrap') or soup.select_one('#main_chart')
    chart_html = ''
    script_text = ''
    if chart_div:
        chart_div_copy = BeautifulSoup(str(chart_div), 'html.parser')
        chart_root = chart_div_copy.select_one('#chartdiv') or chart_div_copy.select_one('.graph-wrap') or chart_div_copy.select_one('#main_chart')
        if chart_root:
            style_val = chart_root.get('style', '')
            if 'height:' not in style_val:
                style_val = (style_val + '; height: 460px;').strip('; ')
                chart_root['style'] = style_val
        chart_html = str(chart_div_copy)

        script_nodes = chart_div.find_all('script')
        if not script_nodes:
            next_script = chart_div.find_next('script')
            if next_script:
                script_nodes.append(next_script)
        if not script_nodes:
            script_nodes = [s for s in soup.find_all('script') if 'main_chart' in (s.get_text('') or '') or 'chartdiv' in (s.get_text('') or '') or 'am4core' in (s.get_text('') or '')]

        if script_nodes:
            script_text = '\n'.join(s.get_text('\n') for s in script_nodes if s)

        def absolutize_urls(html_text: str) -> str:
            html_text = re.sub(r'(["\'])(\/\/(?:images|img)\.badatime\.com[^"\']*)(["\'])', r"https:\1\2\3", html_text)
            html_text = re.sub(r'(["\'])\/img\/icon\/(sunrise|sunset)\.svg(["\'])', r'\1/img/\2.svg\3', html_text)
            return html_text

        def absolutize_script_urls(script_text: str) -> str:
            script_text = re.sub(r'(["\'])\/img\/icon\/(sunrise|sunset)\.svg(["\'])', r'\1/img/\2.svg\3', script_text)
            return script_text

        pc_html = absolutize_urls(pc_html)
        mo_html = absolutize_urls(mo_html)
        chart_html = absolutize_urls(chart_html)
        if script_text:
            script_text = absolutize_script_urls(script_text)

    return success_response({
        'pc_html': pc_html,
        'mo_html': mo_html,
        'chart_html': chart_html,
        'script': script_text,
        'source_url': source_url,
    })

@views.route('/map')
def map_page():
    """지도 페이지 - 항구별 등록된 배 표시"""
    port_coordinates = PORT_COORDINATES
    city_port_mapping = CITY_PORT_MAPPING

    boats = get_all_boats()
    boat_counts = {}
    port_boat_names = {}
    for boat in boats:
        port = boat.port
        if port not in port_boat_names:
            port_boat_names[port] = []
        port_boat_names[port].append(boat.name)

    for boat in boats:
        port = boat.port
        if port in boat_counts:
            boat_counts[port] += 1
        else:
            boat_counts[port] = 1

    total_boats = len(boats)

    return render_template(
        'map.html',
        city_port_mapping=city_port_mapping,
        port_coordinates=port_coordinates,
        boat_counts=boat_counts,
        port_boat_names=port_boat_names,
        total_boats=total_boats
    )

# 추가: 배 삭제 라우트 (POST)
@views.route('/delete/<int:boat_id>', methods=['POST'], endpoint='delete_boat')
def delete_boat_route(boat_id):
    try:
        delete_boat(boat_id)
        flash('배가 삭제되었습니다.', 'success')
    except Exception as e:
        flash(f'삭제 중 오류: {e}', 'danger')
    return redirect(url_for('views.index'))

# New route: handle deletion of selected boats
@views.route('/delete_boats', methods=['POST'])
def delete_boats():
    ids = request.form.getlist('delete_ids')
    if not ids:
        flash('삭제할 배를 선택하세요.', 'warning')
        return redirect(url_for('views.index'))
    deleted = 0
    for bid in ids:
        try:
            # delete_boat 함수가 id를 받는다고 가정
            delete_boat(int(bid))
            deleted += 1
        except Exception as e:
            # continue on error, but notify
            print(f"delete_boat error for id={bid}: {e}")
    flash(f'{deleted}개의 배가 삭제되었습니다.', 'success')
    return redirect(url_for('views.index'))

@views.route('/upload_excel', methods=['POST'])
def upload_excel():
    from models import Boat
    from db import db

    if 'excel_file' not in request.files:
        return jsonify({'success': False, 'message': '파일이 없습니다.'}), 400
    
    file = request.files['excel_file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '파일을 선택해주세요.'}), 400

    if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        try:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            # Map existing boats by name for updates
            existing_by_name = {b.name: b for b in Boat.query.all()}
            
            new_boats_count = 0
            updated_boats_count = 0
            # Iterate over rows, skipping the header (row 1)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Column order from download_excel: No, 지역, 항구, 등록된 배, URL, 비고
                # We ignore 'No' (index 0)
                if len(row) < 5:
                    continue # Skip malformed rows

                city = row[1]
                port = row[2]
                name = row[3]
                url = row[4]
                note = row[5] if len(row) > 5 else None

                # Basic validation
                if not all([city, port, name, url]):
                    current_app.logger.warning(f"Skipping row due to missing data: {row}")
                    continue

                existing = existing_by_name.get(name)
                if not existing:
                    new_boat = Boat(name=name, url=url, city=city, port=port, note=note)
                    db.session.add(new_boat)
                    existing_by_name[name] = new_boat  # track to avoid duplicates
                    new_boats_count += 1
                else:
                    # Update existing boat fields (including 비고)
                    changed = False
                    if existing.city != city and city:
                        existing.city = city; changed = True
                    if existing.port != port and port:
                        existing.port = port; changed = True
                    if existing.url != url and url:
                        existing.url = url; changed = True
                    # note can be empty string; update even if '' provided
                    if note is not None and existing.note != note:
                        existing.note = note; changed = True
                    if changed:
                        updated_boats_count += 1
            
            db.session.commit()
            
            if new_boats_count > 0 or updated_boats_count > 0:
                message = f'성공: 신규 {new_boats_count}척, 업데이트 {updated_boats_count}척 처리했습니다.'
            else:
                message = '변경 사항이 없습니다. 모든 배가 이미 최신입니다.'

            return jsonify({'success': True, 'message': message})

        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Excel upload failed: {e}")
            return jsonify({'success': False, 'message': f'파일 처리 중 오류가 발생했습니다: {e}'}), 500

    return jsonify({'success': False, 'message': '엑셀 파일(.xlsx, .xls)만 업로드할 수 있습니다.'}), 400

# API 엔드포인트: 선박 목록 JSON으로 반환
@views.route('/api/ships', methods=['GET'])
def api_ships():
    """선박 목록을 JSON 형태로 반환하는 API 엔드포인트"""
    try:
        boats = get_all_boats()
        ships_data = []
        
        for boat in boats:
            ship = {
                'id': boat.id,
                'region': boat.city,  # 지역
                'port': boat.port,    # 항구
                'registration_number': boat.name,  # 등록번호 (현재는 name을 사용)
                'name': boat.name,    # 선박 이름
                'url': boat.url       # 상세 URL
            }
            ships_data.append(ship)
        
        return jsonify(ships_data)
    
    except Exception as e:
        current_app.logger.error(f"API ships error: {e}")
        return jsonify({'error': '선박 목록을 가져오는 중 오류가 발생했습니다.'}), 500

# API 엔드포인트: 새 선박 등록
@views.route('/api/ships', methods=['POST'])
def api_add_ship():
    """새 선박을 등록하는 API 엔드포인트"""
    try:
        data = request.get_json()
        
        # 필수 필드 검증
        required_fields = ['region', 'port', 'registrationNumber', 'url']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'{field} 필드가 필요합니다.'}), 400
        
        # 선박 등록
        add_boat_instance(
            name=data.get('registrationNumber'),  # 등록번호를 name으로 사용
            url=data.get('url'),
            city=data.get('region'),
            port=data.get('port')
        )
        
        return jsonify({'success': True, 'message': '선박이 성공적으로 등록되었습니다.'})
        
    except Exception as e:
        current_app.logger.error(f"API add ship error: {e}")
        return jsonify({'error': '선박 등록 중 오류가 발생했습니다.'}), 500

@views.route('/sea-temp-test')
@views.route('/sea-temp-test/<int:port_id>')
def sea_temp_test(port_id=None):
    """Badatime 바다 수온 지도 임베드 테스트 페이지
    - 경로: /sea-temp-test 또는 /sea-temp-test/<port_id>
    - 쿼리스트링으로도 지정 가능: /sea-temp-test?port_id=118
    """
    from services.weather_tide_service import PortDataService

    # 우선순위: path param > query param
    q_port_id = request.args.get('port_id', type=int)
    pid = port_id or q_port_id or 118
    # 항구명-포트ID 매핑 전달하여 선택 편의 제공
    return render_template(
        'sea_temp_test.html',
        current_port_id=pid,
        bada_port_ids=PortDataService.get_bada_port_ids(),
    )

@views.route('/api/sea_temp')
def api_sea_temp():
    """바다타임에서 수온 정보를 가져와서 자체 Kakao Maps API로 재구성"""
    import requests
    from bs4 import BeautifulSoup
    import re
    import json
    
    port_id = request.args.get('port_id', type=int, default=443)
    url = f"https://www.badatime.com/{port_id}/sea-temp"
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }
    
    try:
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # main.content 영역 찾기
        content = soup.select_one('main.content')
        
        if not content:
            return jsonify({'error': 'main.content 영역을 찾을 수 없습니다.'}), 404
        
        # 지도 스크립트에서 마커 데이터 추출
        map_markers = []
        center_coords = {'lat': 34.5049, 'lng': 127.1228}  # 기본값
        
        for script in soup.find_all('script'):
            if script.string and 'mapOption' in script.string:
                script_text = script.string
                
                # 중심 좌표 추출
                center_match = re.search(r'center:\s*new\s+daum\.maps\.LatLng\(([^,]+),\s*([^)]+)\)', script_text)
                if center_match:
                    center_coords = {
                        'lat': float(center_match.group(1)),
                        'lng': float(center_match.group(2))
                    }
                
                # 마커 데이터 추출 (위도, 경도, 라벨, 온도)
                # customOverlay content 추출
                overlay_pattern = r"var content = '(.+?)';\s*var position = new daum\.maps\.LatLng\(([^,]+),\s*([^)]+)\);"
                for match in re.finditer(overlay_pattern, script_text, re.DOTALL):
                    content_html = match.group(1)
                    lat = float(match.group(2))
                    lng = float(match.group(3))
                    
                    # HTML에서 지역명과 온도 추출
                    name_match = re.search(r'font-weight:600[^>]*>([^<]+?)\s*<span', content_html)
                    temp_match = re.search(r'font-size:15px[^>]*>([^<]+)</span>', content_html)
                    time_match = re.search(r'font-size:12px[^>]*>([^<]+)</div>', content_html)
                    
                    if name_match and temp_match:
                        map_markers.append({
                            'name': name_match.group(1).strip(),
                            'temp': temp_match.group(1).strip(),
                            'time': time_match.group(1).strip() if time_match else '',
                            'lat': lat,
                            'lng': lng
                        })
        
        # 이미지 경로를 절대 경로로 변환
        for img in content.find_all('img'):
            src = img.get('src', '')
            if src.startswith('//'):
                img['src'] = 'https:' + src
            elif src.startswith('/'):
                img['src'] = 'https://www.badatime.com' + src
        
        # 링크 경로를 절대 경로로 변환
        for link in content.find_all('a'):
            href = link.get('href', '')
            if href.startswith('/') and not href.startswith('//'):
                link['href'] = 'https://www.badatime.com' + href
                link['target'] = '_blank'
        
        # 지도 관련 스크립트 제거 (우리가 직접 구현)
        for script in content.find_all('script'):
            if 'kakao' in str(script) or 'daum.maps' in str(script) or 'mapContainer' in str(script):
                script.decompose()
        
        # 지도 div는 유지하되 내용 비우기
        map_div = content.find('div', id='map')
        if map_div:
            map_div.clear()
            map_div['style'] = 'width:100%; height:500px;'
        
        # Highcharts 스크립트 추출 (그래프용)
        highcharts_scripts = []
        for script in content.find_all('script'):
            if script.string and 'Highcharts.chart' in script.string:
                highcharts_scripts.append(script.string)
                script.decompose()  # 원본 제거
        
        # content HTML 추출
        content_html = str(content)
        
        return jsonify({
            'success': True,
            'html': content_html,
            'map_data': {
                'center': center_coords,
                'markers': map_markers,
                'zoom': 8
            },
            'highcharts_scripts': highcharts_scripts,
            'source_url': url
        })
        
    except requests.RequestException as e:
        return jsonify({'error': f'요청 중 오류 발생: {str(e)}'}), 500
    except Exception as e:
        return jsonify({'error': f'파싱 중 오류 발생: {str(e)}'}), 500

@views.route('/manifest.json')
def pwa_manifest():
    """Serve the PWA manifest (root scope)."""
    return send_from_directory(current_app.root_path, 'manifest.json', mimetype='application/manifest+json')

@views.route('/service-worker.js')
def pwa_service_worker():
    """Serve the service worker at root for widest scope."""
    return send_from_directory(current_app.root_path, 'service-worker.js', mimetype='application/javascript')

@views.route('/offline')
def offline_page():
    """Offline fallback page served when navigation fails in PWA."""
    return render_template('offline.html')

